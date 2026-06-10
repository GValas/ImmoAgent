"""
scheduler.py — Pipeline continu immo-agent
Tourne indéfiniment, déclenche chaque worker selon sa fréquence propre.

Fréquences (configurables dans criteria.md) :
  Hunter + Analyst   → toutes les N heures  (nouvelles annonces)
  Discovery          → tous les N jours     (re-qualifier les sources)
  Builder            → tous les N jours     (scrapers à jour si sites changent)

Diff :
  Seules les nouvelles annonces (jamais vues) sont scorées et ajoutées
  au fichier de suivi actif data/output/suivi_actif.xlsx.

Usage :
  python scheduler.py            # démarre le scheduler
  python scheduler.py --once     # un seul cycle complet puis stop
"""

import asyncio
import argparse
import json
import re
import hashlib
import builtins as _builtins
import yaml
from datetime import datetime, timedelta
from pathlib import Path

# ── Timestamps automatiques sur tous les prints [Worker] ──────────────────
_orig_print = _builtins.print


def _ts_print(*args, **kwargs):
    if args and isinstance(args[0], str) and args[0].startswith("["):
        ts = datetime.now().strftime("%H:%M:%S")
        _orig_print(f"{ts} {args[0]}", *args[1:], **kwargs)
    else:
        _orig_print(*args, **kwargs)


_builtins.print = _ts_print
# ─────────────────────────────────────────────────────────────────────────

from config_loader import load_criteria, load_sources
from workers import discovery, builder, hunter, analyst

DATA_DIR   = Path("data")
RAW_DIR    = DATA_DIR / "raw"
OUTPUT_DIR = DATA_DIR / "output"
STATE_FILE = DATA_DIR / "scheduler_state.json"
SEEN_FILE  = DATA_DIR / "biens_vus.json"
SUIVI_FILE = OUTPUT_DIR / "suivi_actif.xlsx"

CRITERIA_MD = Path("config/criteria.md")


# ──────────────────────────────────────────────
# CONFIG SCHEDULER DEPUIS criteria.md
# ──────────────────────────────────────────────

def load_scheduler_config() -> dict:
    """Lit les paramètres scheduler depuis criteria.md."""
    defaults = {
        "hunter_interval_hours":   4,
        "discovery_interval_days": 7,
        "builder_interval_days":   30,
        "max_biens_suivi":         50,
    }
    try:
        content = CRITERIA_MD.read_text(encoding="utf-8")
        blocks = re.findall(r"```\s*([\s\S]*?)```", content)
        for block in blocks:
            for line in block.strip().splitlines():
                line = line.split("#")[0].strip()
                if ":" in line:
                    key, val = line.split(":", 1)
                    key = key.strip()
                    if key in defaults:
                        try:
                            defaults[key] = yaml.safe_load(val.strip())
                        except Exception:
                            pass
    except Exception:
        pass
    return defaults


# ──────────────────────────────────────────────
# STATE — dernière exécution de chaque worker
# ──────────────────────────────────────────────

def load_state() -> dict:
    if STATE_FILE.exists():
        return json.loads(STATE_FILE.read_text(encoding="utf-8"))
    return {"last_hunter": None, "last_discovery": None, "last_builder": None}


def save_state(state: dict):
    DATA_DIR.mkdir(exist_ok=True)
    STATE_FILE.write_text(json.dumps(state, indent=2, default=str), encoding="utf-8")


def should_run(last_run_str: str | None, interval_hours: float) -> bool:
    if last_run_str is None:
        return True
    last = datetime.fromisoformat(last_run_str)
    return datetime.now() >= last + timedelta(hours=interval_hours)


# ──────────────────────────────────────────────
# DÉDUPLICATION INTER-RUNS (biens déjà vus)
# ──────────────────────────────────────────────

def load_seen() -> set:
    if SEEN_FILE.exists():
        return set(json.loads(SEEN_FILE.read_text(encoding="utf-8")))
    return set()


def save_seen(seen: set):
    DATA_DIR.mkdir(exist_ok=True)
    SEEN_FILE.write_text(json.dumps(list(seen)), encoding="utf-8")


def bien_hash(b: dict) -> str:
    # ville peut être None (scrapers sans ville exposée : proprietes_rurales,
    # equidomain, horse_immo…) — `or ''` couvre clé absente ET valeur None.
    key = f"{b.get('prix')}-{b.get('surface')}-{str(b.get('ville') or '').lower().strip()}"
    return hashlib.md5(key.encode()).hexdigest()


def bien_identity(b: dict) -> str:
    """Identité stable d'une annonce, INSENSIBLE au prix.

    Sert à dédupliquer le suivi cumulatif : une même annonce dont le vendeur
    a baissé le prix (450 000 → 439 000) doit rester UNE ligne, pas deux.
    Clé sur l'URL normalisée si disponible (identifiant unique de l'annonce),
    sinon repli sur surface+ville (les scrapers sans URL sont rares).
    """
    url = str(b.get("url") or "").strip().lower().rstrip("/")
    if url:
        return "url:" + url
    return "sv:" + f"{b.get('surface')}-{str(b.get('ville') or '').lower().strip()}"


def filter_new(biens: list[dict], seen: set) -> tuple[list[dict], set]:
    """Retourne uniquement les biens jamais vus, et le seen mis à jour."""
    new, updated_seen = [], set(seen)
    for b in biens:
        h = bien_hash(b)
        if h not in updated_seen:
            new.append(b)
            updated_seen.add(h)
    return new, updated_seen


# ──────────────────────────────────────────────
# GARDE-FOU LIVENESS — retire les annonces mortes du suivi
# ──────────────────────────────────────────────

# Marqueurs textuels FORTS d'une annonce retirée (pour les pages qui renvoient 200
# avec un message au lieu d'un 404/410). Volontairement spécifiques pour éviter de
# supprimer une annonce vivante par coïncidence.
_DEAD_MARKERS = (
    "plus disponible", "n'est plus disponible", "annonce supprimée",
    "cette annonce n'existe plus", "annonce n'est plus en ligne",
    "ce bien n'est plus", "no longer available", "property is no longer",
)


async def prune_dead_listings(biens: list[dict], concurrency: int = 8) -> tuple[list[dict], int]:
    """Retire les biens dont l'URL est morte. CONSERVATEUR : on ne retire QUE sur
    404/410/451 ou un marqueur « plus disponible » explicite. Sur timeout, erreur
    réseau, 403 (anti-bot) ou 5xx → on GARDE (pas de suppression sur incertitude).
    Retourne (biens_vivants, nb_retirés)."""
    import httpx
    from urllib.parse import urlparse
    targets = [b for b in biens if str(b.get("url") or "").startswith("http")]
    if not targets:
        return biens, 0
    sem = asyncio.Semaphore(concurrency)
    dead = set()

    async def check(b: dict, client: httpx.AsyncClient):
        url = str(b["url"])
        async with sem:
            try:
                r = await client.get(url, timeout=12)
            except Exception:
                return  # transitoire → garder
            if r.status_code in (404, 410, 451):
                dead.add(id(b))
                return
            if r.status_code != 200:
                return  # 3xx/403/5xx → incertain, garder
            # Redirigé HORS de la fiche (l'identifiant disparaît de l'URL finale)
            # → annonce retirée renvoyée vers un index/accueil (ex. proprietes-privees
            # /annonces/XXX → /annonces-immobilieres). follow_redirects masque le 404.
            req_id = urlparse(url).path.rstrip("/").split("/")[-1]
            if len(req_id) >= 4 and req_id not in urlparse(str(r.url)).path:
                dead.add(id(b))
                return
            low = r.text[:60000].lower()
            if any(m in low for m in _DEAD_MARKERS):
                dead.add(id(b))

    async with httpx.AsyncClient(
        headers={"User-Agent": "Mozilla/5.0 (compatible; immo-agent/1.0)"},
        follow_redirects=True,
    ) as client:
        await asyncio.gather(*[check(b, client) for b in targets])

    alive = [b for b in biens if id(b) not in dead]
    return alive, len(biens) - len(alive)


# ──────────────────────────────────────────────
# MISE À JOUR DU FICHIER DE SUIVI ACTIF
# ──────────────────────────────────────────────

async def update_suivi(new_biens: list[dict], cfg: dict):
    """
    Fusionne les nouveaux biens enrichis dans suivi_actif.xlsx.
    Garde au plus max_biens_suivi biens (triés par match qualitatif).
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    max_biens  = cfg["max_biens_suivi"]

    # Charger le suivi existant
    existing = []
    suivi_json = DATA_DIR / "suivi_actif.json"
    if suivi_json.exists():
        existing = json.loads(suivi_json.read_text(encoding="utf-8"))

    # Scoring retiré → on garde tous les nouveaux biens (à revoir plus tard).
    qualifying = list(new_biens)
    # Les nouveaux d'abord : sur collision d'identité (même annonce), c'est la
    # version fraîche — mieux parsée (géoloc, prix à jour) — qui est conservée.
    merged = qualifying + existing

    # Garde-fou département : ne JAMAIS conserver un bien hors zone cible.
    # Auto-nettoie les fuites historiques du suivi cumulatif (ex. bienici qui
    # renvoyait des biens 05/94 avant le post-filtre _bien_in_dept du scraper).
    try:
        target = {str(d).strip().zfill(2) for d in load_criteria().departements}
    except Exception:
        target = set()
    if target:
        def _dept_ok(b: dict) -> bool:
            cp = str(b.get("code_postal") or "").strip()
            if len(cp) >= 2 and cp[:2].isdigit():
                return cp[:2] in target
            return str(b.get("departement") or "").strip().zfill(2) in target
        before = len(merged)
        merged = [b for b in merged if _dept_ok(b)]
        if before - len(merged):
            print(f"[Scheduler] Garde-fou dept : {before - len(merged)} bien(s) hors-zone écarté(s)")

    # Garde-fou liveness : retire les annonces mortes (vendues/retirées → 404/410
    # ou page « plus disponible »). Le suivi cumulatif les garderait sinon → liens vides.
    try:
        merged, n_dead = await prune_dead_listings(merged)
        if n_dead:
            print(f"[Scheduler] Garde-fou liveness : {n_dead} annonce(s) morte(s) retirée(s)")
    except Exception as e:
        print(f"[Scheduler] Garde-fou liveness ignoré ({e})")

    # « Ajouté le » : date d'entrée dans le suivi. On préserve la date d'origine
    # d'un bien déjà suivi (clé d'identité) ; un bien jamais vu est stampé à
    # aujourd'hui. (Les scrapers ne fournissent pas de date fiable.)
    today = datetime.now().strftime("%Y-%m-%d")
    prior_dates = {bien_identity(b): b.get("date_ajout_suivi")
                   for b in existing if b.get("date_ajout_suivi")}
    for b in merged:
        b["date_ajout_suivi"] = (
            b.get("date_ajout_suivi") or prior_dates.get(bien_identity(b)) or today
        )

    # Déduplique par identité d'annonce (insensible au prix : une baisse de prix
    # ne crée plus de doublon), en gardant la 1ʳᵉ occurrence — donc la version
    # fraîche puisque `qualifying` précède `existing`. PUIS tri + tronque.
    seen_id = set()
    deduped = []
    for b in merged:
        ident = bien_identity(b)
        if ident not in seen_id:
            deduped.append(b)
            seen_id.add(ident)

    # Tri par match qualitatif décroissant (seul signal de pertinence restant).
    deduped.sort(key=lambda x: x.get("match_qualitatif") or 0, reverse=True)
    deduped = deduped[:max_biens]

    # Sauvegarder JSON intermédiaire
    suivi_json.write_text(json.dumps(deduped, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    # Regénérer l'Excel
    _write_suivi_excel(deduped)
    print(f"[Scheduler] Suivi actif : {len(deduped)} biens")


def refilter_suivi():
    """Ré-applique le filtrage A POSTERIORI au suivi cumulatif EXISTANT, sans
    re-scraper, puis régénère data/output/suivi_actif.xlsx.

    Répercute un changement de criteria.md (ex. nouveau mot interdit, nouvelle
    description qualitative) sur le suivi déjà constitué : garde-fou département +
    filtre structurel (prix/surface/pièces/DPE) + mots-clés (obligatoires/interdits)
    + photos_min, puis ré-annotation qualitative NLP (re-score + re-tri), sur les
    données DÉJÀ scrapées et enrichies. Ne re-scrape RIEN (pas de requête liste/
    détail, pas de liveness) ; seul le NLP tourne en local (GPU si dispo)."""
    cfg = load_scheduler_config()
    criteres = load_criteria()
    suivi_json = DATA_DIR / "suivi_actif.json"
    if not suivi_json.exists():
        print(f"[Scheduler] {suivi_json} introuvable — rien à re-filtrer.")
        return

    biens = json.loads(suivi_json.read_text(encoding="utf-8"))
    before = len(biens)

    # Garde-fou département (offline)
    target = {str(d).strip().zfill(2) for d in criteres.departements}
    if target:
        def _dept_ok(b: dict) -> bool:
            cp = str(b.get("code_postal") or "").strip()
            if len(cp) >= 2 and cp[:2].isdigit():
                return cp[:2] in target
            return str(b.get("departement") or "").strip().zfill(2) in target
        biens = [b for b in biens if _dept_ok(b)]

    # Mêmes filtres a posteriori que le Hunter, sur données déjà enrichies.
    biens = hunter.filter_biens(biens, criteres)
    biens = hunter.filter_mots_cles(biens, criteres)
    pmin = getattr(criteres, "photos_min", 0)
    if pmin:
        biens = [b for b in biens if len(b.get("photos") or []) >= pmin]

    # Ré-annotation qualitative (NLP) sur les survivants : répercute un changement
    # de description_qualitative et rafraîchit match_qualitatif avant le tri.
    desc_qual = getattr(criteres, "description_qualitative", "") or ""
    if desc_qual.strip() and biens:
        from workers.qualitative import annotate_biens as qual_annotate
        # refilter_suivi() est synchrone (lancé depuis __main__ hors boucle) :
        # on exécute la coroutine d'annotation via asyncio.run.
        asyncio.run(qual_annotate(biens, desc_qual))

    # Tri par match qualitatif décroissant + plafond max_biens_suivi.
    biens.sort(key=lambda x: x.get("match_qualitatif") or 0, reverse=True)
    biens = biens[:cfg["max_biens_suivi"]]

    suivi_json.write_text(
        json.dumps(biens, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    _write_suivi_excel(biens)
    print(f"[Scheduler] Re-filtrage suivi : {len(biens)}/{before} biens conservés → {SUIVI_FILE}")


def _write_suivi_excel(biens: list[dict]):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[Scheduler] openpyxl manquant — Excel non généré")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Suivi actif"

    # Aligné sur l'export resultats_*.xlsx (workers/analyst.py) + colonne
    # suivi-spécifique « Ajouté le ». Inclut géoloc et liens satellite/cadastre.
    headers = [
        "Match qual.", "Ajouté le", "Source", "Titre",
        "Ville", "Dép", "Département", "Gare", "Bus", "Surface", "Terrain", "Pièces", "DPE",
        "Prix (€)", "Prix/m²", "Prix/m² marché", "Alertes",
        "Satellite", "Ortho+cadastre", "URL"
    ]
    hfill = PatternFill("solid", fgColor="2C3E50")
    hfont = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill
        c.font = hfont
        c.alignment = Alignment(horizontal="center")

    zebra_fill = PatternFill("solid", fgColor="F2F4F4")   # zébrage 1 ligne sur 2

    # Colonnes affichées comme hyperliens : {index_1based: libellé}
    link_labels = {
        headers.index("Satellite") + 1: "Vue satellite",
        headers.index("Ortho+cadastre") + 1: "Ortho + cadastre",
        headers.index("URL") + 1: "Voir l'annonce",
    }
    price_cols = {headers.index(h) + 1 for h in ("Prix (€)", "Prix/m²", "Prix/m² marché", "Terrain")}

    for row, b in enumerate(biens, 2):
        zebra = zebra_fill if row % 2 == 0 else None

        vals  = [
            b.get("match_qualitatif"),
            (b.get("date_ajout_suivi") or "")[:10],
            b.get("source", ""),
            b.get("titre", ""),
            b.get("ville", ""),
            b.get("departement", ""),
            analyst.dept_nom(b.get("departement")),
            (f"{b.get('gare_nom')} ({b.get('gare_distance_km')} km)"
             if b.get("gare_nom") else ""),
            (f"{b.get('bus_nom')} ({b.get('bus_distance_km')} km)"
             if b.get("bus_proche") else ""),
            b.get("surface"),
            b.get("surface_terrain"),
            b.get("pieces"),
            b.get("dpe", ""),
            b.get("prix"),
            b.get("prix_m2_calcule"),
            b.get("prix_m2_marche_dep"),
            " | ".join(b.get("alerte", [])),
            b.get("maps_satellite_url", ""),
            b.get("geoportail_url", ""),
            b.get("url", ""),
        ]
        for col, v in enumerate(vals, 1):
            if isinstance(v, (list, dict)):
                v = str(v) if v else ""
            c = ws.cell(row=row, column=col, value=v)
            # Fond : zébrage 1 ligne sur 2
            if zebra is not None:
                c.fill = zebra
            if col in price_cols and isinstance(v, (int, float)):
                c.number_format = "#,##0"
            if col in link_labels and v:
                c.hyperlink = str(v)
                c.value = link_labels[col]
                c.style = "Hyperlink"

    widths = [8, 12, 12, 40, 18, 6, 18, 24, 22, 9, 9, 8, 6, 12, 10, 14, 35,
              20, 14, 16, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws2 = wb.create_sheet("Infos")
    ws2["A1"] = f"Dernière mise à jour : {ts}"
    ws2["A2"] = f"Nombre de biens suivis : {len(biens)}"

    wb.save(SUIVI_FILE)
    print(f"[Scheduler] Excel mis à jour → {SUIVI_FILE}")


# ──────────────────────────────────────────────
# CYCLE PRINCIPAL
# ──────────────────────────────────────────────

async def run_cycle(state: dict, cfg: dict, sources: list[dict]) -> dict:
    """Exécute un cycle du scheduler selon l'état et les fréquences."""
    criteres = load_criteria()
    now      = datetime.now().isoformat()
    ran_something = False

    # ── Discovery ──
    if should_run(state["last_discovery"], cfg["discovery_interval_days"] * 24):
        print(f"\n[Scheduler] {_ts()} Discovery...")
        sources = await discovery.run(criteres)
        state["last_discovery"] = now
        ran_something = True

    # ── Builder ──
    if should_run(state["last_builder"], cfg["builder_interval_days"] * 24):
        print(f"[Scheduler] {_ts()} Builder...")
        await builder.run(sources, criteres)
        state["last_builder"] = now
        ran_something = True

    # ── Hunter + Analyst ──
    if should_run(state["last_hunter"], cfg["hunter_interval_hours"]):
        print(f"[Scheduler] {_ts()} Hunter...")
        biens_bruts = await hunter.run(sources, criteres)

        if biens_bruts:
            # Filtrer les biens déjà vus
            seen = load_seen()
            new_biens, seen = filter_new(biens_bruts, seen)
            save_seen(seen)

            print(f"[Scheduler] {len(new_biens)} nouveau(x) bien(s) sur {len(biens_bruts)}")

            if new_biens:
                print(f"[Scheduler] {_ts()} Analyst sur {len(new_biens)} nouveaux biens...")
                from workers.analyst import enrich_bien, fetch_prix_marche_dvf
                prix_marche = await fetch_prix_marche_dvf(criteres.departements)

                # Match qualitatif NLP (si une description est définie)
                desc_qual = getattr(criteres, "description_qualitative", "") or ""
                if desc_qual.strip():
                    from workers.qualitative import annotate_biens as qual_annotate
                    await qual_annotate(new_biens, desc_qual)

                enriched = [enrich_bien(b, prix_marche) for b in new_biens]
                await update_suivi(enriched, cfg)
            else:
                print("[Scheduler] Aucun nouveau bien — garde-fous sur suivi existant")
                # Pas de nouveau bien, mais on repasse le suivi aux garde-fous
                # (ban/dept) : une image ban ajoutée entre deux runs purge alors
                # les entrées déjà accumulées.
                await update_suivi([], cfg)
        else:
            print("[Scheduler] Aucun bien récupéré ce cycle")

        state["last_hunter"] = now
        ran_something = True

    if not ran_something:
        print(f"[Scheduler] {_ts()} Rien à faire ce tick")

    return state


def _ts() -> str:
    return datetime.now().strftime("%H:%M:%S")


def _next_run(state: dict, cfg: dict) -> datetime:
    """Calcule le prochain moment où quelque chose doit tourner."""
    candidates = []
    if state["last_hunter"]:
        candidates.append(
            datetime.fromisoformat(state["last_hunter"])
            + timedelta(hours=cfg["hunter_interval_hours"])
        )
    else:
        candidates.append(datetime.now())
    return min(candidates) if candidates else datetime.now()


# ──────────────────────────────────────────────
# BOUCLE INFINIE
# ──────────────────────────────────────────────

async def run_forever():
    print("=" * 55)
    print("  IMMO-AGENT Scheduler — démarrage")
    print("=" * 55)

    DATA_DIR.mkdir(exist_ok=True)
    state   = load_state()
    cfg     = load_scheduler_config()
    sources = load_sources()

    print(f"  Hunter    : toutes les {cfg['hunter_interval_hours']}h")
    print(f"  Discovery : tous les {cfg['discovery_interval_days']}j")
    print(f"  Builder   : tous les {cfg['builder_interval_days']}j")
    print("=" * 55)

    while True:
        cfg   = load_scheduler_config()   # relit criteria.md à chaque cycle
        state = await run_cycle(state, cfg, sources)
        save_state(state)

        # Calcul du prochain tick
        next_run = _next_run(state, cfg)
        wait_sec = max(0, (next_run - datetime.now()).total_seconds())
        print(f"[Scheduler] Prochain cycle dans {int(wait_sec // 3600)}h{int((wait_sec % 3600) // 60)}m")
        await asyncio.sleep(wait_sec)


async def run_once():
    """Un seul cycle complet — utile pour tester."""
    DATA_DIR.mkdir(exist_ok=True)
    state   = {"last_hunter": None, "last_discovery": None, "last_builder": None}
    cfg     = load_scheduler_config()
    sources = load_sources()
    state   = await run_cycle(state, cfg, sources)
    save_state(state)
    print("\n[Scheduler] Cycle unique terminé.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--once", action="store_true", help="Un seul cycle puis stop")
    parser.add_argument("--refilter", action="store_true",
                        help="Ré-applique le filtrage a posteriori au suivi cumulatif "
                             "existant et régénère suivi_actif.xlsx, sans re-scraper")
    args = parser.parse_args()

    if args.refilter:
        refilter_suivi()
    elif args.once:
        asyncio.run(run_once())
    else:
        asyncio.run(run_forever())
