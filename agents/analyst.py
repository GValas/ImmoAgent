"""
agents/analyst.py — Agent 4 : Analyst
Score chaque bien, enrichit avec données DVF réelles, agrège dans un Excel final.
"""
import json
import asyncio
from datetime import datetime
from pathlib import Path

# Import DVF scraper pour les prix de référence réels (CSV data.gouv.fr)
import sys as _sys
_sys.path.insert(0, str(Path(__file__).parent.parent))
from scrapers.dvf import get_prix_m2_reference as _dvf_get_prix_m2

OUTPUT_DIR = Path(__file__).parent.parent / "data" / "output"

# ──────────────────────────────────────────────
# SCORING
# ──────────────────────────────────────────────

DPE_SCORE = {"A": 100, "B": 85, "C": 70, "D": 55, "E": 30, "F": 10, "G": 0}

def score_bien(bien: dict, criteres_dict: dict, prix_m2_marche: dict) -> dict:
    """
    Calcule un score pondéré sur 100 pour un bien.
    Retourne le bien enrichi avec score_total et score_detail.
    """
    poids = criteres_dict["poids_scoring"]
    scores = {}
    alertes = []

    # --- Prix vs marché ---
    dep = bien.get("departement", "")
    prix_m2_ref = prix_m2_marche.get(dep, 0)
    prix_m2_bien = bien.get("prix_m2") or (
        bien["prix"] / bien["surface"] if bien.get("prix") and bien.get("surface") else None
    )
    if prix_m2_bien and prix_m2_ref:
        ratio = prix_m2_ref / prix_m2_bien  # >1 = bonne affaire
        scores["prix"] = min(100, max(0, int(ratio * 50 + 50)))
        if ratio > 1.3:
            alertes.append("🟢 Prix très inférieur au marché")
        elif ratio < 0.8:
            alertes.append("🔴 Prix supérieur au marché")
    else:
        scores["prix"] = 50  # neutre si pas de données

    # --- Surface ---
    surface = bien.get("surface", 0) or 0
    s_min = criteres_dict.get("surface_min", 80)
    s_max = criteres_dict.get("surface_max", 300)
    if surface >= s_max:
        scores["surface"] = 100
    elif surface >= s_min:
        scores["surface"] = int((surface - s_min) / (s_max - s_min) * 100)
    else:
        scores["surface"] = 0

    # --- Terrain ---
    terrain = bien.get("surface_terrain", 0) or 0
    t_min = criteres_dict.get("terrain_min", 200)
    scores["terrain"] = min(100, int(terrain / max(t_min, 1) * 60)) if terrain else 20

    # --- DPE ---
    dpe = (bien.get("dpe") or "").upper()
    scores["dpe"] = DPE_SCORE.get(dpe, 50)
    if dpe in ["F", "G"]:
        alertes.append(f"⚠️ DPE {dpe} — passoire thermique")

    # --- Localisation (proxy : photos disponibles + adresse) ---
    has_photos = len(bien.get("photos", [])) > 0
    has_address = bool(bien.get("adresse") or bien.get("ville"))
    scores["localisation"] = (50 if has_address else 20) + (30 if has_photos else 0)

    # --- État (proxy : mots-clés dans description) ---
    desc = ((bien.get("description") or "") + " " + (bien.get("titre") or "")).lower()
    if any(w in desc for w in ["neuf", "rénov", "refait", "récent"]):
        scores["etat"] = 90
    elif any(w in desc for w in ["travaux", "rénover", "à rafraîchir"]):
        scores["etat"] = 30
        alertes.append("🔧 Travaux probables")
    else:
        scores["etat"] = 60

    # --- Score total pondéré ---
    # Les clés poids_scoring ont le préfixe "poids_"
    total = sum(
        scores.get(k, 50) * poids.get(f"poids_{k}", poids.get(k, 0)) / 100
        for k in ["prix", "surface", "terrain", "localisation", "etat", "dpe"]
    )

    bien["score_total"] = round(total, 1)
    bien["score_detail"] = scores
    bien["alerte"] = alertes
    bien["prix_m2_calcule"] = round(prix_m2_bien, 0) if prix_m2_bien else None
    bien["prix_m2_marche_dep"] = prix_m2_ref or None

    return bien


# ──────────────────────────────────────────────
# DVF — prix marché par département
# ──────────────────────────────────────────────

async def fetch_prix_marche_dvf(departements: list[str]) -> dict:
    """
    Récupère le prix médian €/m² réel par département via les fichiers CSV DVF (data.gouv.fr).
    Télécharge les données de transactions passées (maisons 2024) et calcule la médiane.
    Retourne {dep_code: prix_m2_median}.
    """
    try:
        result = await _dvf_get_prix_m2(departements)
        print(f"[Analyst] Prix DVF réels chargés pour {len(result)} depts")
        return result
    except Exception as e:
        print(f"[Analyst] DVF indisponible ({e}) — fallback valeurs de référence")
        return {dep: _prix_m2_reference(dep) for dep in departements}


def _prix_m2_reference(dep: str) -> int:
    """
    Prix de référence €/m² maison par département — fallback si DVF indisponible.
    Valeurs estimées 2024 pour les départements cibles (Loire, Centre, Sarthe).
    """
    refs = {
        # Départements cibles — Loire / Centre-Val-de-Loire / Pays-de-la-Loire
        "72": 1650,  # Sarthe
        "28": 1950,  # Eure-et-Loir
        "45": 1900,  # Loiret
        "89": 1400,  # Yonne
        "49": 2100,  # Maine-et-Loire
        "37": 2200,  # Indre-et-Loire
        "36": 1200,  # Indre
        "18": 1300,  # Cher
        "58": 1150,  # Nièvre
        "41": 1800,  # Loir-et-Cher
        "53": 1600,  # Mayenne
        # Autres grandes zones
        "06": 5200, "83": 4100, "84": 2800, "13": 3500,
        "69": 4800, "75": 10500, "92": 7200, "94": 5100,
        "33": 3800, "34": 3600, "44": 3700, "31": 3400,
        "67": 3200, "76": 2600, "59": 2400, "38": 3000,
    }
    return refs.get(dep, 2000)  # défaut national réaliste (non plus 3000)


# ──────────────────────────────────────────────
# RÉSUMÉ LLM
# ──────────────────────────────────────────────

def llm_summary(top_biens: list[dict]) -> str:
    """
    Génère un résumé texte local des meilleures opportunités.
    Sans appel API — pour un résumé enrichi, demande à Claude Code :
    "Analyse data/output/resultats_xxx.xlsx et résume les meilleures opportunités"
    """
    if not top_biens:
        return "Aucun bien scoré disponible."

    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    lignes = [f"Résumé généré le {ts} — {len(top_biens)} bien(s) analysé(s)\n"]

    for i, b in enumerate(top_biens[:5], 1):
        score  = b.get("score_total", 0)
        prix   = f"{b.get('prix', '?'):,} €" if b.get("prix") else "prix ?"
        surf   = f"{b.get('surface', '?')} m²"
        ville  = b.get("ville", "?")
        dep    = b.get("departement", "")
        dpe    = b.get("dpe", "?")
        els    = b.get("elements_detectes") or []
        el_str = (" | ⚠️ " + ", ".join(e["nom"] for e in els)) if els else ""
        alertes = b.get("alerte", [])
        alert_str = f" ⚠ {alertes[0]}" if alertes else ""
        lignes.append(
            f"#{i} [{score:.0f}/100{el_str}] {b.get('titre','')[:50]}\n"
            f"   {ville} ({dep}) — {surf} — {prix} — DPE {dpe}{alert_str}"
        )

    return "\n".join(lignes)


# ──────────────────────────────────────────────
# EXPORT EXCEL
# ──────────────────────────────────────────────

# Département (code INSEE) → nom en toutes lettres. Partagé avec scheduler.py.
DEPT_NOMS = {
    "01": "Ain", "02": "Aisne", "03": "Allier", "04": "Alpes-de-Haute-Provence",
    "05": "Hautes-Alpes", "06": "Alpes-Maritimes", "07": "Ardèche", "08": "Ardennes",
    "09": "Ariège", "10": "Aube", "11": "Aude", "12": "Aveyron",
    "13": "Bouches-du-Rhône", "14": "Calvados", "15": "Cantal", "16": "Charente",
    "17": "Charente-Maritime", "18": "Cher", "19": "Corrèze", "2A": "Corse-du-Sud",
    "2B": "Haute-Corse", "21": "Côte-d'Or", "22": "Côtes-d'Armor", "23": "Creuse",
    "24": "Dordogne", "25": "Doubs", "26": "Drôme", "27": "Eure", "28": "Eure-et-Loir",
    "29": "Finistère", "30": "Gard", "31": "Haute-Garonne", "32": "Gers",
    "33": "Gironde", "34": "Hérault", "35": "Ille-et-Vilaine", "36": "Indre",
    "37": "Indre-et-Loire", "38": "Isère", "39": "Jura", "40": "Landes",
    "41": "Loir-et-Cher", "42": "Loire", "43": "Haute-Loire", "44": "Loire-Atlantique",
    "45": "Loiret", "46": "Lot", "47": "Lot-et-Garonne", "48": "Lozère",
    "49": "Maine-et-Loire", "50": "Manche", "51": "Marne", "52": "Haute-Marne",
    "53": "Mayenne", "54": "Meurthe-et-Moselle", "55": "Meuse", "56": "Morbihan",
    "57": "Moselle", "58": "Nièvre", "59": "Nord", "60": "Oise", "61": "Orne",
    "62": "Pas-de-Calais", "63": "Puy-de-Dôme", "64": "Pyrénées-Atlantiques",
    "65": "Hautes-Pyrénées", "66": "Pyrénées-Orientales", "67": "Bas-Rhin",
    "68": "Haut-Rhin", "69": "Rhône", "70": "Haute-Saône", "71": "Saône-et-Loire",
    "72": "Sarthe", "73": "Savoie", "74": "Haute-Savoie", "75": "Paris",
    "76": "Seine-Maritime", "77": "Seine-et-Marne", "78": "Yvelines",
    "79": "Deux-Sèvres", "80": "Somme", "81": "Tarn", "82": "Tarn-et-Garonne",
    "83": "Var", "84": "Vaucluse", "85": "Vendée", "86": "Vienne",
    "87": "Haute-Vienne", "88": "Vosges", "89": "Yonne", "90": "Territoire de Belfort",
    "91": "Essonne", "92": "Hauts-de-Seine", "93": "Seine-Saint-Denis",
    "94": "Val-de-Marne", "95": "Val-d'Oise", "971": "Guadeloupe", "972": "Martinique",
    "973": "Guyane", "974": "La Réunion", "976": "Mayotte",
}


def dept_nom(code) -> str:
    """Nom du département en toutes lettres ('72' → 'Sarthe'). Repli : le code brut."""
    if code is None:
        return ""
    return DEPT_NOMS.get(str(code).strip().zfill(2), str(code))


def export_excel(biens: list[dict], resume: str) -> Path:
    """Exporte les résultats scorés dans un fichier Excel."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import numbers as xl_numbers
    except ImportError:
        raise ImportError("pip install openpyxl")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    path = OUTPUT_DIR / f"resultats_{ts}.xlsx"

    wb = openpyxl.Workbook()

    # ── Feuille 1 : Résultats ──
    ws = wb.active
    ws.title = "Résultats"

    headers = [
        "Score", "Source", "Titre", "Ville",
        "Dép", "Département", "Gare", "Bus",
        "Surface", "Terrain", "Pièces", "DPE",
        "Prix (€)", "Prix/m²", "Prix/m² marché",
        "Résumé vision", "Alertes", "Piscine hors-sol", "Éléments détectés",
        "Parcelle probable", "Satellite", "Ortho+cadastre", "URL"
    ]
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Couleur du Score (qualité, sur la seule cellule Score) + zébrage des lignes.
    score_fills = {
        "high":   PatternFill("solid", fgColor="D5F5E3"),  # vert
        "medium": PatternFill("solid", fgColor="FEF9E7"),  # jaune
        "low":    PatternFill("solid", fgColor="FADBD8"),  # rouge
    }
    zebra_fill = PatternFill("solid", fgColor="F2F4F4")    # 1 ligne sur 2

    # Colonnes affichées comme hyperliens : {index_1based: libellé du lien}
    link_labels = {
        headers.index("Satellite") + 1: "Vue satellite",
        headers.index("Ortho+cadastre") + 1: "Ortho + cadastre",
        headers.index("URL") + 1: "Voir l'annonce",
    }
    score_col = headers.index("Score") + 1
    price_cols = {headers.index(h) + 1 for h in ("Prix (€)", "Prix/m²", "Prix/m² marché", "Terrain")}

    for row, b in enumerate(biens, 2):
        score = b.get("score_total", 0)
        score_fill = (score_fills["high"] if score >= 70
                      else score_fills["medium"] if score >= 45 else score_fills["low"])
        zebra = zebra_fill if row % 2 == 0 else None   # 1 ligne sur 2

        values = [
            score,
            b.get("source", ""),
            b.get("titre", ""),
            b.get("ville", ""),
            b.get("departement", ""),
            dept_nom(b.get("departement")),
            (f"{b.get('gare_nom')} ({b.get('gare_distance_km')} km)"
             if b.get("gare_nom") else ""),
            (f"{b.get('bus_nom')} ({b.get('bus_distance_km')} km)"
             if b.get("bus_proche") else ""),
            b.get("surface"),
            b.get("surface_terrain"),
            b.get("pieces"),
            b.get("dpe", ""),
            b.get("prix"),
            b.get("prix_m2_calcule"),
            b.get("prix_m2_marche_dep"),
            b.get("resume_visuel", ""),
            " | ".join(a for a in b.get("alerte", []) if "piscine_hors_sol" not in a),
            next((f"🛟 {e['score']}" for e in (b.get("elements_detectes") or [])
                  if e["nom"] == "piscine_hors_sol"), ""),
            " | ".join(
                f"{'⛔' if e.get('mode') == 'exclusion' else '⚠️'} {e['nom']} ({e['score']})"
                for e in (b.get("elements_detectes") or [])
                if e["nom"] != "piscine_hors_sol"
            ),
            b.get("parcelle_match", ""),
            b.get("maps_satellite_url", ""),
            b.get("geoportail_url", ""),
            b.get("url", ""),
        ]
        for col, val in enumerate(values, 1):
            if isinstance(val, (list, dict)):
                val = str(val) if val else ""
            cell = ws.cell(row=row, column=col, value=val)
            # Fond : Score en couleur de qualité, le reste en zébrage 1 ligne/2
            if col == score_col:
                cell.fill = score_fill
            elif zebra is not None:
                cell.fill = zebra
            # Séparateur de milliers sur les prix
            if col in price_cols and isinstance(val, (int, float)):
                cell.number_format = "#,##0"
            if col in link_labels and val:
                cell.hyperlink = str(val)
                cell.value = link_labels[col]
                cell.style = "Hyperlink"  # style natif Excel → change de couleur après clic

    # Largeurs colonnes
    widths = [8, 12, 40, 20, 6, 18, 24, 22, 9, 9, 8, 6, 12, 10, 14, 45, 40,
              14, 26, 20, 14, 16, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # ── Feuille 2 : Résumé ──
    ws2 = wb.create_sheet("Résumé")
    ws2["A1"] = "Résumé exécutif"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A3"] = resume
    ws2["A3"].alignment = Alignment(wrap_text=True)
    ws2.column_dimensions["A"].width = 100

    wb.save(path)
    print(f"[Analyst] Excel exporté → {path}")
    return path


# ──────────────────────────────────────────────
# POINT D'ENTRÉE
# ──────────────────────────────────────────────

async def run(biens_bruts: list[dict], criteres) -> Path:
    """Pipeline complet : enrichissement DVF → scoring → export Excel."""
    print(f"[Analyst] Scoring de {len(biens_bruts)} biens...")

    criteres_dict = {
        "surface_min": criteres.surface_min,
        "surface_max": criteres.surface_max,
        "terrain_min": criteres.terrain_min,
        "poids_scoring": criteres.poids_scoring,
    }

    # Données marché DVF
    prix_marche = await fetch_prix_marche_dvf(criteres.departements)

    # Scoring
    scored = [score_bien(b, criteres_dict, prix_marche) for b in biens_bruts]

    # Tri par score décroissant
    scored.sort(key=lambda b: b.get("score_total", 0), reverse=True)

    print(f"[Analyst] Top 5 scores : {[b['score_total'] for b in scored[:5]]}")

    # Résumé LLM
    resume = llm_summary(scored[:10])
    print(f"\n--- RÉSUMÉ EXÉCUTIF ---\n{resume}\n")

    # Export
    path = export_excel(scored, resume)
    return path


if __name__ == "__main__":
    import sys
    from pathlib import Path
    from config_loader import load_criteria

    raw_files = sorted(Path("data/raw").glob("*.json"))
    if not raw_files:
        print("Aucun fichier raw trouvé. Lance d'abord hunter.py")
        sys.exit(1)

    latest = raw_files[-1]
    print(f"Analyse de {latest}")
    biens = json.loads(latest.read_text(encoding="utf-8"))
    criteres = load_criteria()
    asyncio.run(run(biens, criteres))
