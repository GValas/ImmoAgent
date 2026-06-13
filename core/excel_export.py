"""core/excel_export.py — Écriture unique des classeurs Excel.

Auparavant deux writers quasi identiques (analyst.export_excel et
scheduler._write_suivi_excel) divergeaient (tableaux de largeurs désynchronisés
des en-têtes). On définit ici UN registre de colonnes (en-tête → extracteur +
largeur) et un writer unique ; chaque appelant ne fournit que la LISTE ordonnée
des en-têtes voulus, ce qui rend tout décalage largeur/colonne impossible.
"""
from __future__ import annotations

from collections.abc import Callable
from pathlib import Path
from typing import Optional

from core.dept_data import dept_nom


def _accessibilite(b: dict) -> str:
    from scrapers.geolocate import rome2rio_url  # import tardif (évite cycle au chargement)
    return b.get("rome2rio_url") or rome2rio_url(b.get("ville", ""), b.get("code_postal"))


# Registre : en-tête → (extracteur(bien) -> valeur, largeur de colonne).
# Toute colonne affichée par un classeur doit figurer ici.
COLUMN_REGISTRY: dict[str, tuple[Callable[[dict], object], int]] = {
    "Match qual.":    (lambda b: b.get("match_qualitatif"), 8),
    "Ajouté le":      (lambda b: (b.get("date_ajout_suivi") or "")[:10], 12),
    "Source":         (lambda b: b.get("source", ""), 12),
    "Titre":          (lambda b: b.get("titre", ""), 40),
    "Ville":          (lambda b: b.get("ville", ""), 20),
    "Dép":            (lambda b: b.get("departement", ""), 6),
    "Département":     (lambda b: dept_nom(b.get("departement")), 18),
    "Gare":           (lambda b: f"{b.get('gare_nom')} ({b.get('gare_distance_km')} km)"
                       if b.get("gare_nom") else "", 24),
    "Bus":            (lambda b: f"{b.get('bus_nom')} ({b.get('bus_distance_km')} km)"
                       if b.get("bus_proche") else "", 22),
    "Accessibilité":  (_accessibilite, 16),
    "Surface":        (lambda b: b.get("surface"), 9),
    "Terrain":        (lambda b: b.get("surface_terrain"), 9),
    "Pièces":         (lambda b: b.get("pieces"), 8),
    "DPE":            (lambda b: b.get("dpe", ""), 6),
    "Prix (€)":       (lambda b: b.get("prix"), 12),
    "Prix/m²":        (lambda b: b.get("prix_m2_calcule"), 10),
    "Prix/m² marché": (lambda b: b.get("prix_m2_marche_dep"), 14),
    "Alertes":        (lambda b: " | ".join(b.get("alerte", [])), 35),
    "Extrait qual.":  (lambda b: b.get("match_extrait", ""), 40),
    "Satellite":      (lambda b: b.get("maps_satellite_url", ""), 20),
    "Ortho+cadastre": (lambda b: b.get("geoportail_url", ""), 14),
    "URL":            (lambda b: b.get("url", ""), 16),
}

# En-têtes rendus comme hyperliens (libellé affiché à la place de l'URL).
LINK_LABELS = {
    "Accessibilité": "Paris ▸ train",
    "Satellite": "Vue satellite",
    "Ortho+cadastre": "Ortho + cadastre",
    "URL": "Voir l'annonce",
}
# En-têtes formatés en nombre avec séparateur de milliers.
PRICE_HEADERS = {"Prix (€)", "Prix/m²", "Prix/m² marché", "Terrain"}

# Jeux de colonnes des deux classeurs (même registre, ordre/colonnes propres à chacun).
RESULTATS_COLUMNS = [
    "Match qual.", "Source", "Titre", "Ville", "Dép", "Département", "Gare", "Bus",
    "Accessibilité", "Surface", "Terrain", "Pièces", "DPE", "Prix (€)", "Prix/m²",
    "Prix/m² marché", "Alertes", "Extrait qual.", "Satellite", "Ortho+cadastre", "URL",
]
SUIVI_COLUMNS = [
    "Match qual.", "Ajouté le", "Source", "Titre", "Ville", "Dép", "Département",
    "Gare", "Bus", "Accessibilité", "Surface", "Terrain", "Pièces", "DPE", "Prix (€)",
    "Prix/m²", "Prix/m² marché", "Alertes", "Satellite", "Ortho+cadastre", "URL",
]


def write_listings_xlsx(
    biens: list[dict],
    path: Path,
    *,
    columns: list[str],
    sheet_title: str,
    build_extra_sheet: Optional[Callable[[object], None]] = None,
) -> Path:
    """Écrit un classeur de biens à `path` avec les `columns` (en-têtes du registre).

    `build_extra_sheet(workbook)` permet d'ajouter une 2ᵉ feuille (résumé, infos).
    Lève ImportError si openpyxl est absent — l'appelant gère (analyst lève,
    scheduler journalise)."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(columns, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    zebra_fill = PatternFill("solid", fgColor="F2F4F4")   # zébrage 1 ligne sur 2
    link_cols = {i + 1: LINK_LABELS[h] for i, h in enumerate(columns) if h in LINK_LABELS}
    price_cols = {i + 1 for i, h in enumerate(columns) if h in PRICE_HEADERS}
    extractors = [COLUMN_REGISTRY[h][0] for h in columns]

    for row, b in enumerate(biens, 2):
        zebra = zebra_fill if row % 2 == 0 else None
        for col, extract in enumerate(extractors, 1):
            v = extract(b)
            if isinstance(v, (list, dict)):
                v = str(v) if v else ""
            c = ws.cell(row=row, column=col, value=v)
            if zebra is not None:
                c.fill = zebra
            if col in price_cols and isinstance(v, (int, float)):
                c.number_format = "#,##0"
            if col in link_cols and v:
                c.hyperlink = str(v)
                c.value = link_cols[col]
                c.style = "Hyperlink"

    for col, h in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col)].width = COLUMN_REGISTRY[h][1]

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    if build_extra_sheet is not None:
        build_extra_sheet(wb)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
