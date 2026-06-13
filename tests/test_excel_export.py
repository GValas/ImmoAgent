"""Tests du writer Excel unique (core.excel_export)."""
import openpyxl

from core.excel_export import (
    COLUMN_REGISTRY,
    RESULTATS_COLUMNS,
    SUIVI_COLUMNS,
    write_listings_xlsx,
)

SAMPLE = [
    {"source": "bienici", "titre": "Maison Tours", "ville": "Tours",
     "departement": "37", "surface": 120, "surface_terrain": 800, "pieces": 5,
     "dpe": "D", "prix": 295000, "prix_m2_calcule": 2458, "match_qualitatif": 82,
     "url": "https://ex.fr/1", "code_postal": "37000"},
    {"source": "iad", "titre": "Longère", "ville": "Le Mans",
     "departement": "72", "surface": 160, "prix": 320000, "url": "https://ex.fr/2",
     "code_postal": "72000"},
]


def test_toutes_les_colonnes_sont_dans_le_registre():
    # Garde-fou : aucune colonne déclarée sans extracteur/largeur (cause des
    # décalages largeur/colonne de l'ancien code).
    for h in set(RESULTATS_COLUMNS) | set(SUIVI_COLUMNS):
        assert h in COLUMN_REGISTRY


def test_write_resultats_xlsx(tmp_path):
    path = tmp_path / "resultats.xlsx"
    write_listings_xlsx(SAMPLE, path, columns=RESULTATS_COLUMNS, sheet_title="Résultats")
    wb = openpyxl.load_workbook(path)
    ws = wb["Résultats"]
    # En-têtes alignés et nombre de lignes = 1 (header) + len(biens).
    headers = [c.value for c in ws[1]]
    assert headers == RESULTATS_COLUMNS
    assert ws.max_row == 1 + len(SAMPLE)


def test_write_suivi_xlsx_avec_feuille_extra(tmp_path):
    path = tmp_path / "suivi.xlsx"

    def _infos(wb):
        ws2 = wb.create_sheet("Infos")
        ws2["A1"] = "Test"

    write_listings_xlsx(SAMPLE, path, columns=SUIVI_COLUMNS,
                        sheet_title="Suivi actif", build_extra_sheet=_infos)
    wb = openpyxl.load_workbook(path)
    assert wb["Suivi actif"][1][0].value == "Match qual."
    assert "Ajouté le" in [c.value for c in wb["Suivi actif"][1]]
    assert wb["Infos"]["A1"].value == "Test"
